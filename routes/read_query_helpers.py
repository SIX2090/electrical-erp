"""Read/query helpers: CSV/Excel export, pagination, and query parameter parsing for list pages."""
import csv
import io
import math
from datetime import date, datetime
from decimal import Decimal

from flask import Response, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


_column_cache = {}
_safe_rows = None
_safe_one = None


def configure_read_query_helpers(safe_rows, safe_one):
    global _safe_rows, _safe_one
    _safe_rows = safe_rows
    _safe_one = safe_one


def clear_column_cache():
    _column_cache.clear()


def _require_safe_rows():
    if _safe_rows is None:
        raise RuntimeError("read query helpers are not configured")
    return _safe_rows


def _require_safe_one():
    if _safe_one is None:
        raise RuntimeError("read query helpers are not configured")
    return _safe_one


def _columns(*items):
    return [{"key": key, "label": label} for key, label in items]


def _table_columns(table):
    if table in _column_cache:
        return _column_cache[table]
    rows = _require_safe_rows()(
        """
        SELECT column_name, data_type
        FROM information_schema.columns
        WHERE table_schema='public' AND table_name=%s
        ORDER BY ordinal_position
        """,
        (table,),
    )
    _column_cache[table] = rows
    return rows


def _has_table(table):
    return bool(_table_columns(table))


def _list_columns(table, preferred=None):
    table_cols = [row["column_name"] for row in _table_columns(table)]
    labels = {
        "id": "ID",
        "code": "\u7f16\u7801",
        "name": "\u540d\u79f0",
        "specification": "\u89c4\u683c\u578b\u53f7",
        "unit": "\u5355\u4f4d",
        "contact_person": "\u8054\u7cfb\u4eba",
        "phone": "\u7535\u8bdd",
        "customer_level": "\u5ba2\u6237\u7b49\u7ea7",
        "warehouse_id": "\u4ed3\u5e93",
        "location_id": "\u5e93\u4f4d",
        "product_id": "\u7269\u6599",
        "supplier_id": "\u4f9b\u5e94\u5546",
        "customer_id": "\u5ba2\u6237",
        "order_no": "\u8ba2\u5355\u53f7",
        "request_no": "\u7533\u8bf7\u5355\u53f7",
        "receipt_no": "\u6536\u6b3e/\u6536\u8d27\u5355\u53f7",
        "payment_no": "\u4ed8\u6b3e\u5355\u53f7",
        "shipment_no": "\u53d1\u8d27\u5355\u53f7",
        "return_no": "\u9000\u8d27\u5355\u53f7",
        "invoice_no": "\u53d1\u7968\u53f7",
        "quote_no": "\u62a5\u4ef7\u5355\u53f7",
        "wo_no": "\u5de5\u5355\u53f7",
        "bom_no": "BOM\u7f16\u53f7",
        "project_code": "\u9879\u76ee\u53f7",
        "cabinet_no": "\u673a\u53f7",
        "status": "\u72b6\u6001",
        "quantity": "\u6570\u91cf",
        "total_amount": "\u603b\u91d1\u989d",
        "received_amount": "\u5df2\u6536\u91d1\u989d",
        "paid_amount": "\u5df2\u4ed8\u91d1\u989d",
        "amount": "\u91d1\u989d",
        "balance": "\u4f59\u989d",
        "created_at": "\u521b\u5efa\u65f6\u95f4",
        "remark": "\u5907\u6ce8",
    }
    if preferred:
        keys = [key for key in preferred if key in table_cols] if table_cols else list(preferred)
    else:
        priority = [
            "id",
            "code",
            "name",
            "order_no",
            "request_no",
            "receipt_no",
            "shipment_no",
            "wo_no",
            "bom_no",
            "project_code",
            "cabinet_no",
            "status",
            "quantity",
            "total_amount",
            "amount",
            "balance",
            "created_at",
            "remark",
        ]
        keys = [key for key in priority if key in table_cols]
        for key in table_cols:
            if key not in keys and len(keys) < 8:
                keys.append(key)
    return [{"key": key, "label": labels.get(key, key)} for key in keys[:10]]


def _text_columns(table):
    return [
        row["column_name"]
        for row in _table_columns(table)
        if row["data_type"] in {"character varying", "text", "character"}
    ][:8]


def _quote_identifier(value):
    return '"' + str(value).replace('"', '""') + '"'


def _first_existing(table_cols, candidates):
    for candidate in candidates:
        if candidate in table_cols:
            return candidate
    return None


def _safe_int_arg(name, default, minimum, maximum):
    try:
        value = int(request.args.get(name) or default)
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


def _select_rows(table, preferred=None, limit=100):
    if not _has_table(table):
        return [], _list_columns(table, preferred)
    table_cols = [row["column_name"] for row in _table_columns(table)]
    table_types = {row["column_name"]: row["data_type"] for row in _table_columns(table)}
    params = []
    where_parts = []
    keyword = (request.args.get("keyword") or request.args.get("q") or request.args.get("search") or "").strip()
    if keyword:
        text_cols = _text_columns(table)
        if text_cols:
            where_parts.append("(" + " OR ".join(f"{_quote_identifier(col)} ILIKE %s" for col in text_cols) + ")")
            params.extend([f"%{keyword}%"] * len(text_cols))
    status = (request.args.get("status") or "").strip()
    if status and "status" in table_cols:
        where_parts.append(f"COALESCE({_quote_identifier('status')}, '') ILIKE %s")
        params.append(f"%{status}%")
    project = (request.args.get("project") or request.args.get("project_cabinet") or "").strip()
    if project:
        project_cols = [col for col in ("project_code", "project_no", "cabinet_no", "cabinet_no") if col in table_cols]
        if project_cols:
            where_parts.append("(" + " OR ".join(f"COALESCE({_quote_identifier(col)}, '') ILIKE %s" for col in project_cols) + ")")
            params.extend([f"%{project}%"] * len(project_cols))
    date_col = _first_existing(
        table_cols,
        ("doc_date", "order_date", "request_date", "receipt_date", "shipment_date", "completion_date", "report_date", "created_at", "updated_at"),
    )
    date_from = (request.args.get("date_from") or request.args.get("date_start") or "").strip()
    date_to = (request.args.get("date_to") or request.args.get("date_end") or "").strip()
    if date_col and date_from:
        where_parts.append(f"{_quote_identifier(date_col)} >= %s")
        params.append(date_from)
    if date_col and date_to:
        where_parts.append(f"{_quote_identifier(date_col)} <= %s")
        params.append(date_to)
    allowed_sort_cols = set(table_cols)
    requested_sort = (request.args.get("sort") or "").strip()
    default_sort = _first_existing(
        table_cols,
        ("doc_date", "order_date", "request_date", "receipt_date", "shipment_date", "completion_date", "report_date", "updated_at", "created_at", "id"),
    ) or table_cols[0]
    sort_col = requested_sort if requested_sort in allowed_sort_cols else default_sort
    direction = "ASC" if (request.args.get("direction") or "").lower() == "asc" else "DESC"
    nulls = "NULLS LAST" if direction == "ASC" else "NULLS FIRST"
    order_parts = [f"{_quote_identifier(sort_col)} {direction} {nulls}"]
    if sort_col != "id" and "id" in table_cols:
        order_parts.append(f"{_quote_identifier('id')} DESC")
    page_size = _safe_int_arg("page_size", limit, 20, 500)
    page = _safe_int_arg("page", 1, 1, 100000)
    offset = (page - 1) * page_size
    where = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    rows = _require_safe_rows()(
        f"SELECT * FROM {_quote_identifier(table)}{where} ORDER BY {', '.join(order_parts)} LIMIT %s OFFSET %s",
        tuple(params + [page_size, offset]),
    )
    total_row = _require_safe_one()(f"SELECT COUNT(*) AS total FROM {_quote_identifier(table)}{where}", tuple(params))
    total = int((total_row or {}).get("total") or 0)
    request.list_query_meta = {
        "sort": sort_col,
        "direction": direction.lower(),
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": max(1, int(math.ceil(total / page_size))) if page_size else 1,
        "date_column": date_col,
    }
    return rows, _list_columns(table, preferred)


def _export_format():
    value = (request.args.get("format") or request.args.get("export") or "csv").strip().lower()
    return "xlsx" if value in {"xlsx", "excel"} else "csv"


def _safe_excel_sheet_title(filename):
    title = "".join(ch for ch in str(filename or "export") if ch not in r"[]:*?/\\").strip() or "export"
    return title[:31]


def _excel_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date, int, float, bool)):
        return value
    return str(value)


def _xlsx_response(rows, filename):
    rows = list(rows or [])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_excel_sheet_title(filename)
    if rows:
        if hasattr(rows[0], "keys"):
            keys = list(rows[0].keys())
            sheet.append(keys)
            for row in rows:
                sheet.append([_excel_cell_value(row.get(key)) for key in keys])
            width_source = [[key] + [row.get(key, "") for row in rows[:100]] for key in keys]
        else:
            for row in rows:
                sheet.append([_excel_cell_value(value) for value in row])
            width_source = []
            max_columns = max((len(row) for row in rows), default=0)
            for index in range(max_columns):
                width_source.append([row[index] if index < len(row) else "" for row in rows[:100]])
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for index, values in enumerate(width_source, start=1):
            values = [str(value) for value in values]
            width = min(max(len(value) for value in values) + 2, 42)
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


def _csv_response(rows, filename):
    if _export_format() == "xlsx":
        return _xlsx_response(rows, filename)
    output = io.StringIO()
    rows = list(rows or [])
    writer = csv.writer(output)
    if rows:
        keys = list(rows[0].keys())
        writer.writerow(keys)
        for row in rows:
            writer.writerow([row.get(key, "") for key in keys])
    return Response(
        output.getvalue().encode("utf-8-sig"),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
    )


def _safe_table_ref(table):
    """Quote table identifier only when it is a plain identifier (no JOIN/space).
    Complex FROM clauses (e.g. 'a JOIN b ON ...') are passed through as-is."""
    text = str(table or "")
    if " " in text or " join " in text.lower():
        return text
    return _quote_identifier(text)


def _safe_column_ref(column):
    """Quote column identifier only when it is a plain identifier (no function/space)."""
    text = str(column or "")
    if " " in text or "(" in text:
        return text
    return _quote_identifier(text)


def _count_rows(table, where="", params=None):
    clause = f" WHERE {where}" if where else ""
    row = _require_safe_one()(f"SELECT COUNT(*) AS value FROM {_safe_table_ref(table)}{clause}", params or ())
    return row.get("value", 0) if row else 0


def _sum_value(table, column, where="", params=None):
    clause = f" WHERE {where}" if where else ""
    row = _require_safe_one()(f"SELECT COALESCE(SUM({_safe_column_ref(column)}), 0) AS value FROM {_safe_table_ref(table)}{clause}", params or ())
    return row.get("value", 0) if row else 0


def _has_column(table, column):
    return any(row.get("column_name") == column for row in _table_columns(table))
